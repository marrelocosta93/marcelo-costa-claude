#!/usr/bin/env python3
"""
xlsx_components.py — Componentes reutilizáveis para geração de planilhas XLSX.
Equivalente ao pptx_components.py para PowerPoint.

Usa openpyxl para read/write e xlsxwriter para geração otimizada.

Uso:
    from xlsx_components import (
        create_workbook, add_data_sheet, add_summary_sheet,
        style_headers, apply_zebra, auto_fit_columns,
        add_formula_row, add_chart, validate_formulas
    )
"""

from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, Protection, numbers
    )
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False


# ============================================================
# PALETAS DE CORES CORPORATIVAS
# ============================================================

PALETTES = {
    "corporate_blue": {
        "primary": "2F5496",
        "secondary": "4472C4",
        "accent": "ED7D31",
        "light": "D6E4F0",
        "text": "333333",
        "header_bg": "2F5496",
        "header_fg": "FFFFFF",
        "zebra_light": "F2F2F2",
        "zebra_dark": "D9E2F3",
        "positive": "C6EFCE",
        "negative": "FFC7CE",
        "neutral": "FFEB9C",
    },
    "modern_gray": {
        "primary": "404040",
        "secondary": "808080",
        "accent": "00B050",
        "light": "F2F2F2",
        "text": "262626",
        "header_bg": "404040",
        "header_fg": "FFFFFF",
        "zebra_light": "F8F8F8",
        "zebra_dark": "E8E8E8",
        "positive": "C6EFCE",
        "negative": "FFC7CE",
        "neutral": "FFEB9C",
    },
    "raiz_brand": {
        "primary": "F08700",
        "secondary": "7AC5BF",
        "accent": "2F5496",
        "light": "FFF3E0",
        "text": "333333",
        "header_bg": "F08700",
        "header_fg": "FFFFFF",
        "zebra_light": "FFF8F0",
        "zebra_dark": "FFE8CC",
        "positive": "C6EFCE",
        "negative": "FFC7CE",
        "neutral": "FFEB9C",
    },
}

# ============================================================
# FORMATOS NUMÉRICOS BR
# ============================================================

FORMATS_BR = {
    "currency": 'R$ #,##0.00',
    "percent": '0.0%',
    "date": 'DD/MM/YYYY',
    "datetime": 'DD/MM/YYYY HH:MM',
    "integer": '#,##0',
    "decimal": '#,##0.00',
    "text": '@',
}


# ============================================================
# FUNÇÕES OPENPYXL
# ============================================================

def create_workbook(title: str = None, palette: str = "corporate_blue") -> "openpyxl.Workbook":
    """Cria workbook com configuração padrão."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl não instalado: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title or "Dados"

    # Margem coluna A
    ws.column_dimensions['A'].width = 3

    return wb


def setup_sheet(ws, title: str = None, freeze: bool = True, tab_color: str = None):
    """Configura uma sheet com padrões: freeze panes, margem coluna A."""
    if title:
        ws.title = title

    # Margem coluna A
    ws.column_dimensions['A'].width = 3

    # Freeze panes (header na linha 2)
    if freeze:
        ws.freeze_panes = 'B3'

    # Tab color
    if tab_color:
        ws.sheet_properties.tabColor = tab_color


def style_headers(ws, row: int = 2, palette: str = "corporate_blue",
                  start_col: int = 2, end_col: int = None):
    """Aplica estilo profissional aos headers."""
    colors = PALETTES.get(palette, PALETTES["corporate_blue"])

    header_font = Font(name='Calibri', size=11, bold=True, color=colors["header_fg"])
    header_fill = PatternFill(start_color=colors["header_bg"],
                              end_color=colors["header_bg"], fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    if end_col is None:
        end_col = ws.max_column

    for col_idx in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def apply_zebra(ws, start_row: int = 3, end_row: int = None,
                start_col: int = 2, end_col: int = None,
                palette: str = "corporate_blue"):
    """Aplica linhas alternadas (zebra stripes)."""
    colors = PALETTES.get(palette, PALETTES["corporate_blue"])
    light_fill = PatternFill(start_color=colors["zebra_light"],
                             end_color=colors["zebra_light"], fill_type='solid')

    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = light_fill


def auto_fit_columns(ws, min_width: int = 8, max_width: int = 50, padding: int = 2):
    """Auto-ajusta largura das colunas baseado no conteúdo."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        # Pular coluna A (margem)
        if col_letter == 'A':
            continue

        for cell in col:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len

        adjusted_width = max(min_width, min(max_length + padding, max_width))
        ws.column_dimensions[col_letter].width = adjusted_width


def add_formula_row(ws, row: int, formulas: dict[str, str],
                    font_bold: bool = True, palette: str = "corporate_blue"):
    """Adiciona linha de totais/fórmulas.

    Args:
        ws: Worksheet
        row: Número da linha
        formulas: Dict de {coluna: fórmula} ex: {"C": "=SUM(C3:C50)"}
        font_bold: Negrito
        palette: Paleta de cores
    """
    colors = PALETTES.get(palette, PALETTES["corporate_blue"])
    total_fill = PatternFill(start_color=colors["light"],
                             end_color=colors["light"], fill_type='solid')
    total_font = Font(name='Calibri', size=11, bold=font_bold, color=colors["text"])

    for col_letter, formula in formulas.items():
        cell = ws[f"{col_letter}{row}"]
        cell.value = formula
        cell.font = total_font
        cell.fill = total_fill


def add_conditional_formatting(ws, range_str: str,
                               palette: str = "corporate_blue"):
    """Adiciona formatação condicional: verde > 0, vermelho < 0."""
    colors = PALETTES.get(palette, PALETTES["corporate_blue"])

    ws.conditional_formatting.add(range_str,
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=PatternFill(start_color=colors["positive"],
                                    end_color=colors["positive"], fill_type='solid')))
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill(start_color=colors["negative"],
                                    end_color=colors["negative"], fill_type='solid')))


def add_data_validation_list(ws, range_str: str, options: list[str],
                              allow_blank: bool = True):
    """Adiciona validação de dados com lista suspensa."""
    formula = '"' + ','.join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = "Valor inválido. Selecione da lista."
    dv.errorTitle = "Erro de Validação"
    ws.add_data_validation(dv)
    dv.add(range_str)


def add_excel_table(ws, ref: str, name: str, style: str = "TableStyleMedium2"):
    """Adiciona tabela Excel formatada."""
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tab)


def add_bar_chart(ws, title: str, data_ref: tuple, cats_ref: tuple,
                  anchor: str = "F3", width: int = 20, height: int = 12):
    """Adiciona gráfico de barras.

    Args:
        data_ref: (min_col, min_row, max_col, max_row)
        cats_ref: (min_col, min_row, max_row)
    """
    chart = BarChart()
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height

    data = Reference(ws, min_col=data_ref[0], min_row=data_ref[1],
                     max_col=data_ref[2], max_row=data_ref[3])
    cats = Reference(ws, min_col=cats_ref[0], min_row=cats_ref[1], max_row=cats_ref[2])

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)
    return chart


def add_line_chart(ws, title: str, data_ref: tuple, cats_ref: tuple,
                   anchor: str = "F3", width: int = 20, height: int = 12):
    """Adiciona gráfico de linhas."""
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height

    data = Reference(ws, min_col=data_ref[0], min_row=data_ref[1],
                     max_col=data_ref[2], max_row=data_ref[3])
    cats = Reference(ws, min_col=cats_ref[0], min_row=cats_ref[1], max_row=cats_ref[2])

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)
    return chart


def add_pie_chart(ws, title: str, data_ref: tuple, cats_ref: tuple,
                  anchor: str = "F3", width: int = 15, height: int = 12):
    """Adiciona gráfico de pizza."""
    chart = PieChart()
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height

    data = Reference(ws, min_col=data_ref[0], min_row=data_ref[1],
                     max_col=data_ref[2], max_row=data_ref[3])
    cats = Reference(ws, min_col=cats_ref[0], min_row=cats_ref[1], max_row=cats_ref[2])

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)
    return chart


def setup_print(ws, area: str = None, title_rows: str = '2:2',
                landscape: bool = True):
    """Configura impressão."""
    if area:
        ws.print_area = area
    ws.print_title_rows = title_rows
    ws.page_setup.orientation = (
        ws.ORIENTATION_LANDSCAPE if landscape else ws.ORIENTATION_PORTRAIT
    )
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def format_column(ws, col_letter: str, number_format: str,
                  start_row: int = 3, end_row: int = None):
    """Aplica formato numérico a uma coluna inteira."""
    fmt = FORMATS_BR.get(number_format, number_format)
    if end_row is None:
        end_row = ws.max_row

    for row_idx in range(start_row, end_row + 1):
        ws[f"{col_letter}{row_idx}"].number_format = fmt


# ============================================================
# VALIDAÇÃO
# ============================================================

def validate_formulas(filepath: str) -> list[str]:
    """Valida que não há erros de fórmula no workbook."""
    if not HAS_OPENPYXL:
        return ["openpyxl não instalado"]

    issues = []
    wb = openpyxl.load_workbook(filepath, data_only=True)
    error_patterns = ['#REF!', '#NAME?', '#VALUE!', '#DIV/0!', '#NULL!', '#N/A']

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for ep in error_patterns:
                        if ep in str(cell.value):
                            issues.append(
                                f"Aba '{ws.title}' célula {cell.coordinate}: {ep}"
                            )

    # Verificar se freeze panes existe
    for ws in wb.worksheets:
        if ws.freeze_panes is None:
            issues.append(f"Aba '{ws.title}': Freeze panes não configurado")

    return issues


def validate_structure(filepath: str) -> list[str]:
    """Valida estrutura padrão do workbook."""
    if not HAS_OPENPYXL:
        return ["openpyxl não instalado"]

    issues = []
    wb = openpyxl.load_workbook(filepath)

    for ws in wb.worksheets:
        # Verificar coluna A como margem
        col_a_has_data = False
        for row in ws.iter_rows(min_col=1, max_col=1, min_row=3):
            for cell in row:
                if cell.value and str(cell.value).strip():
                    col_a_has_data = True
                    break
        if col_a_has_data:
            issues.append(f"Aba '{ws.title}': Coluna A contém dados (deveria ser margem)")

        # Verificar headers na linha 2
        has_header = False
        for cell in ws[2]:
            if cell.value:
                has_header = True
                break
        if not has_header and ws.max_row > 2:
            issues.append(f"Aba '{ws.title}': Sem headers na linha 2")

    return issues


# ============================================================
# QUICK BUILDERS
# ============================================================

def quick_data_sheet(wb, title: str, headers: list[str],
                     data: list[list[Any]], palette: str = "corporate_blue",
                     tab_color: str = None) -> Any:
    """Cria sheet completa com dados, headers, zebra, auto-fit."""
    ws = wb.create_sheet(title=title)
    setup_sheet(ws, title=title, tab_color=tab_color)

    # Headers na linha 2 (col B em diante)
    for col_idx, header in enumerate(headers, start=2):
        ws.cell(row=2, column=col_idx, value=header)

    # Dados a partir da linha 3
    for row_idx, row_data in enumerate(data, start=3):
        for col_idx, value in enumerate(row_data, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Estilizar
    style_headers(ws, row=2, palette=palette, start_col=2, end_col=len(headers) + 1)
    apply_zebra(ws, start_row=3, start_col=2, end_col=len(headers) + 1, palette=palette)
    auto_fit_columns(ws)

    # AutoFilter
    end_col_letter = get_column_letter(len(headers) + 1)
    ws.auto_filter.ref = f"B2:{end_col_letter}{len(data) + 2}"

    return ws


def quick_summary_sheet(wb, title: str = "Resumo",
                        kpis: dict[str, Any] = None,
                        palette: str = "corporate_blue") -> Any:
    """Cria aba de resumo com KPIs."""
    ws = wb.create_sheet(title=title, index=0)  # Primeira aba
    setup_sheet(ws, title=title, freeze=False,
                tab_color=PALETTES.get(palette, PALETTES["corporate_blue"])["primary"])

    colors = PALETTES.get(palette, PALETTES["corporate_blue"])

    if kpis:
        # Título
        ws.cell(row=2, column=2, value="Indicador")
        ws.cell(row=2, column=3, value="Valor")
        style_headers(ws, row=2, palette=palette, start_col=2, end_col=3)

        for row_idx, (key, value) in enumerate(kpis.items(), start=3):
            label_cell = ws.cell(row=row_idx, column=2, value=key)
            label_cell.font = Font(name='Calibri', size=11, bold=True, color=colors["text"])

            value_cell = ws.cell(row=row_idx, column=3, value=value)
            value_cell.font = Font(name='Calibri', size=14, bold=True, color=colors["primary"])
            value_cell.alignment = Alignment(horizontal='center')

        apply_zebra(ws, start_row=3, end_row=len(kpis) + 2,
                    start_col=2, end_col=3, palette=palette)
        auto_fit_columns(ws)

    return ws
